import uuid
import csv
import io
from pathlib import Path

from fastapi import APIRouter, Request, Depends, Form, UploadFile, File, BackgroundTasks
from fastapi.responses import RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session, selectinload, joinedload

from app.database import get_db
from app.models import Product, LabelVersion, ComplianceCheck
from app.routes.auth import get_current_user_from_cookie
from app.routes.labels import _process_label  # single shared scan pipeline
from app.config import get_settings
from app.utils.alerts import get_unread_alert_count

router = APIRouter()
settings = get_settings()
templates = Jinja2Templates(directory=str(Path(__file__).parent.parent / "templates"))

LABEL_EXTENSIONS = {".jpg", ".jpeg", ".png", ".tiff", ".tif", ".pdf", ".webp"}
SPREADSHEET_EXTENSIONS = {".csv", ".xlsx", ".xls"}
ALL_ALLOWED = LABEL_EXTENSIONS | SPREADSHEET_EXTENSIONS


def require_user(request: Request, db: Session):
    user = get_current_user_from_cookie(request, db)
    if not user:
        return None
    return user


@router.get("/products/export.csv")
async def export_products_csv(request: Request, db: Session = Depends(get_db)):
    from fastapi.responses import Response as FastResponse
    from app.models import CheckResult

    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    products = (
        db.query(Product)
        .filter(Product.user_id == user.id, Product.is_active == True, Product.is_quick_check == False)
        .order_by(Product.created_at.desc())
        .all()
    )
    for p in products:
        for lv in p.label_versions:
            _ = lv.checks

    output = io.StringIO()
    w = csv.writer(output)
    w.writerow(["Product Name", "SKU", "Category", "Latest Score", "Status",
                "Last Scanned", "Critical Failures", "High Failures", "Top Violation"])
    for p in products:
        lbl = p.latest_label
        score = p.compliance_score
        status = "Compliant" if score is not None and score >= 80 else ("Flagged" if score is not None else "Not Scanned")
        last_scan = lbl.uploaded_at.strftime("%Y-%m-%d") if lbl else ""
        crit = high = 0
        top_violation = ""
        if lbl and lbl.checks:
            failed = [c for c in lbl.checks if c.result == CheckResult.FAIL]
            crit = sum(1 for c in failed if c.rule and c.rule.severity.value == "CRITICAL")
            high = sum(1 for c in failed if c.rule and c.rule.severity.value == "HIGH")
            if failed:
                sev_order = ["CRITICAL", "HIGH", "MEDIUM", "LOW"]
                worst = sorted(
                    failed,
                    key=lambda c: sev_order.index(c.rule.severity.value if c.rule and c.rule.severity else "LOW")
                )[0]
                top_violation = (worst.rule.description[:80] if worst.rule else "")
        w.writerow([p.name, p.sku or "", p.category or "", score if score is not None else "",
                    status, last_scan, crit, high, top_violation])

    return FastResponse(
        output.getvalue(), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=products_compliance.csv"},
    )


@router.get("/products")
async def products_list(request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    products = (
        db.query(Product)
        .filter(
            Product.user_id == user.id,
            Product.is_active == True,
            Product.is_quick_check == False,  # hide ephemeral checker products
        )
        .order_by(Product.created_at.desc())
        .all()
    )
    for p in products:
        _ = p.label_versions
        for lv in p.label_versions:
            _ = lv.checks

    archived_count = (
        db.query(Product)
        .filter(Product.user_id == user.id, Product.is_active == False)
        .count()
    )

    return templates.TemplateResponse("products.html", {
        "request": request,
        "user": user,
        "products": products,
        "archived_count": archived_count,
        "flash_message": request.query_params.get("msg"),
        "flash_type": request.query_params.get("type", "info"),
    })


@router.post("/products/add")
async def add_product(
    request: Request,
    background_tasks: BackgroundTasks,
    name: str = Form(...),
    sku: str = Form(""),
    category: str = Form("Health Supplement"),
    description: str = Form(""),
    file: UploadFile = File(None),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Quota check — fail open on internal error but log so it can be detected
    from app.services.quota_service import check_product_limit
    import logging as _log
    try:
        quota_allowed, quota_msg = check_product_limit(user, db)
    except Exception as _qe:
        _log.getLogger(__name__).error("[quota] Product quota check failed for user %s: %s", user.id, _qe)
        quota_allowed, quota_msg = True, ""
    if not quota_allowed:
        from urllib.parse import quote
        return RedirectResponse(url=f"/products?msg={quote(quota_msg)}&type=error", status_code=302)

    # ── If a CSV/Excel file was uploaded, handle as bulk import ──
    if file and file.filename:
        suffix = Path(file.filename).suffix.lower()

        if suffix in SPREADSHEET_EXTENSIONS:
            return await _handle_spreadsheet_upload(request, user, file, suffix, db)

        if suffix not in LABEL_EXTENSIONS:
            from urllib.parse import quote
            msg = f"Unsupported file type '{suffix}'. Use image/PDF for labels or CSV/Excel for bulk import."
            return RedirectResponse(url=f"/products?msg={quote(msg)}&type=error", status_code=302)

    # ── Create the product ──
    product = Product(
        user_id=user.id,
        name=name,
        sku=sku or None,
        category=category,
        description=description,
    )
    db.add(product)
    db.commit()

    # ── If an image/PDF was uploaded, attach as label and trigger processing ──
    if file and file.filename:
        suffix = Path(file.filename).suffix.lower()
        if suffix in LABEL_EXTENSIONS:
            label_version = await _save_label_file(file, suffix, product.id, db)
            background_tasks.add_task(_process_label, label_version.id)
            try:
                from app.services.activity_service import log_action
                log_action(user.id, "label_uploaded", "label", label_version.id,
                           detail=f"Uploaded label for {product.name}")
            except Exception:
                pass
            return RedirectResponse(url=f"/labels/{label_version.id}?processing=1", status_code=302)

    return RedirectResponse(url=f"/products/{product.id}", status_code=302)


async def _save_label_file(file: UploadFile, suffix: str, product_id: int, db: Session) -> LabelVersion:
    """Save uploaded file to disk and create a LabelVersion record."""
    content = await file.read()

    # Reject files over 50 MB
    if len(content) > 50 * 1024 * 1024:
        raise ValueError("File too large (max 50 MB)")

    upload_dir = Path(settings.upload_dir) / str(product_id)
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_name = f"{uuid.uuid4().hex}{suffix}"
    file_path = upload_dir / file_name

    with open(file_path, "wb") as f:
        f.write(content)

    # Mark previous versions as not current
    db.query(LabelVersion).filter(
        LabelVersion.product_id == product_id, LabelVersion.is_current == True
    ).update({"is_current": False})

    label_version = LabelVersion(
        product_id=product_id,
        file_path=str(file_path),
        file_name=file.filename,
        file_type="pdf" if suffix == ".pdf" else "image",
        is_current=True,
        file_data=content,
    )
    db.add(label_version)
    db.commit()
    db.refresh(label_version)
    return label_version


# _process_label_bg removed — now uses shared _process_label from labels.py
# This ensures /products upload and /labels upload use the exact same pipeline
# with _SCAN_JOBS tracking, proper error recovery, and correct API signatures.


def _feed_product_to_llm(product_id: int, db):
    """Ingest/update this product's data + extraction into the Products LLM KB.

    Stores the full extraction JSON, compliance results, and OCR text so the
    LLM can reference all learned data without making additional API calls.
    This is the 'learning' step — every scan result is permanently stored.
    """
    import json as _json
    from app.models import KBDocument, KBChunk, KBType, Product, LabelVersion, ComplianceCheck, ComplianceRule, CheckResult
    from app.services.llm_service import _ingest_document

    product = db.query(Product).filter(Product.id == product_id).first()
    if not product:
        return
    # Don't pollute the product KB with ephemeral quick-check runs
    if getattr(product, 'is_quick_check', False):
        return

    # Remove any existing KB document for this product
    existing = db.query(KBDocument).filter(
        KBDocument.kb_type == KBType.PRODUCTS,
        KBDocument.source.like(f"db:product:{product.id}%"),
    ).all()
    for doc in existing:
        db.query(KBChunk).filter(KBChunk.document_id == doc.id).delete()
        db.delete(doc)
    db.flush()

    # Build content
    latest_lv = (
        db.query(LabelVersion)
        .filter(LabelVersion.product_id == product.id, LabelVersion.is_current == True)
        .first()
    )

    checks_summary = ""
    extraction_text = ""
    if latest_lv:
        checks = (
            db.query(ComplianceCheck)
            .filter(ComplianceCheck.label_version_id == latest_lv.id)
            .all()
        )
        total = len(checks)
        passed = sum(1 for c in checks if c.result == CheckResult.PASS)
        failed = [c for c in checks if c.result == CheckResult.FAIL]

        fail_lines = []
        for fc in failed[:10]:
            rule = db.query(ComplianceRule).filter(ComplianceRule.id == fc.rule_id).first()
            rule_code = rule.rule_code if rule else "Unknown"
            fail_lines.append(f"  - FAIL [{rule_code}]: {fc.message or 'No detail'}")

        from app.services.compliance_engine import calculate_compliance_score as _calc_score
        score = _calc_score(checks)
        checks_summary = (
            f"\nLabel Analysis (uploaded {latest_lv.uploaded_at.strftime('%Y-%m-%d')}):\n"
            f"Compliance Score: {score}% ({passed}/{total} checks passed, severity-weighted)\n"
            f"Failing Checks:\n"
            + ("\n".join(fail_lines) if fail_lines else "  None")
            + f"\nOCR Text Preview: {(latest_lv.ocr_raw_text or '')[:400]}"
        )

        # Store full extraction data so the LLM learns from every scan
        if latest_lv.extraction_json:
            try:
                extraction_text = (
                    f"\n\nExtracted Label Data (confidence: {latest_lv.extraction_confidence or 0:.0%}):\n"
                    + _json.dumps(latest_lv.extraction_json, indent=2, ensure_ascii=False)[:3000]
                )
            except Exception:
                pass

    content = (
        f"Product: {product.name}\n"
        f"SKU: {product.sku or 'N/A'}\n"
        f"Category: {product.category or 'Nutraceutical'}\n"
        f"Description: {product.description or 'N/A'}\n"
        f"Created: {product.created_at.strftime('%Y-%m-%d')}"
        + checks_summary
        + extraction_text
    )

    _ingest_document(
        db, "products",
        title=f"Product: {product.name}",
        source=f"db:product:{product.id}",
        content=content,
    )


async def _handle_spreadsheet_upload(request: Request, user, file: UploadFile, suffix: str, db: Session):
    """Parse CSV or Excel file as bulk product import."""
    from urllib.parse import quote
    from app.services.quota_service import check_product_limit, get_product_headroom

    # Quota gate — check before any work
    allowed, quota_msg = check_product_limit(user, db)
    if not allowed:
        return RedirectResponse(url=f"/products?msg={quote(quota_msg)}&type=error", status_code=302)

    content_bytes = await file.read()
    rows = []

    try:
        if suffix == ".csv":
            text = content_bytes.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            for row in reader:
                rows.append(row)
        elif suffix in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    d = {}
                    for i, val in enumerate(row):
                        if i < len(headers) and headers[i]:
                            d[headers[i]] = str(val).strip() if val is not None else ""
                    if d.get("name"):
                        rows.append(d)
                wb.close()
            except ImportError:
                msg = "Excel support requires openpyxl. Use CSV format instead."
                return RedirectResponse(url=f"/products?msg={quote(msg)}&type=error", status_code=302)
    except Exception as e:
        msg = f"Failed to parse file: {str(e)[:100]}"
        return RedirectResponse(url=f"/products?msg={quote(msg)}&type=error", status_code=302)

    if not rows:
        msg = "No data rows found in file. Ensure headers include: name, sku, category, description"
        return RedirectResponse(url=f"/products?msg={quote(msg)}&type=error", status_code=302)

    # Cap rows to remaining quota headroom
    headroom = get_product_headroom(user, db)
    truncated = False
    if headroom is not None and len(rows) > headroom:
        truncated = True
        rows = rows[:headroom]

    VALID_CATS = {
        "health supplement", "nutraceutical", "functional food",
        "food for special dietary use", "novel food", "ayurvedic / asu",
    }

    added = 0
    for row in rows:
        name = row.get("name", "").strip()
        if not name:
            continue
        sku = row.get("sku", "").strip()
        cat = row.get("category", "Health Supplement").strip()
        desc = row.get("description", "").strip()

        if cat.lower() not in VALID_CATS:
            cat = "Health Supplement"

        existing = db.query(Product).filter(
            Product.user_id == user.id,
            Product.name == name,
            Product.is_active == True,
        ).first()
        if existing:
            continue

        p = Product(user_id=user.id, name=name, sku=sku or None, category=cat, description=desc)
        db.add(p)
        added += 1

    db.commit()
    msg = f"Imported {added} product(s) from {file.filename}"
    if truncated:
        msg += " — some rows were skipped because you've reached your plan's product limit. Upgrade to import more."
    msg_type = "success" if added > 0 else "info"
    return RedirectResponse(url=f"/products?msg={quote(msg)}&type={msg_type}", status_code=302)


@router.get("/products/bulk-upload")
async def bulk_upload_page(request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    from app.models import Alert, AlertStatus
    unread_alerts = get_unread_alert_count(user, db)

    return templates.TemplateResponse("bulk_upload.html", {
        "request": request,
        "user": user,
        "unread_alerts": unread_alerts,
        "result": None,
    })


@router.post("/products/bulk-upload")
async def bulk_upload_post(
    request: Request,
    csv_data: str = Form(...),
    db: Session = Depends(get_db),
):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Quota gate
    from app.services.quota_service import check_product_limit, get_product_headroom
    from urllib.parse import quote as _quote
    allowed, quota_msg = check_product_limit(user, db)
    if not allowed:
        return RedirectResponse(url=f"/products?msg={_quote(quota_msg)}&type=error", status_code=302)

    lines = [l.strip() for l in csv_data.strip().splitlines() if l.strip()]
    added, skipped, errors = [], [], []

    # Skip header if present
    if lines and lines[0].lower().startswith("name"):
        lines = lines[1:]

    # Cap to remaining quota headroom
    headroom = get_product_headroom(user, db)
    if headroom is not None and len(lines) > headroom:
        errors.append(f"Import capped at {headroom} row(s) — plan limit reached. Upgrade to import more.")
        lines = lines[:headroom]

    VALID_CATS = {
        "health supplement", "nutraceutical", "functional food",
        "food for special dietary use", "novel food", "ayurvedic / asu",
    }

    for i, line in enumerate(lines, 1):
        parts = [p.strip().strip('"') for p in line.split(",")]
        if not parts or not parts[0]:
            continue
        name = parts[0]
        sku  = parts[1] if len(parts) > 1 else ""
        cat  = parts[2] if len(parts) > 2 else "Health Supplement"
        desc = parts[3] if len(parts) > 3 else ""

        if cat.lower() not in VALID_CATS:
            cat = "Health Supplement"

        # Skip duplicates
        existing = db.query(Product).filter(
            Product.user_id == user.id,
            Product.name == name,
            Product.is_active == True,
        ).first()
        if existing:
            skipped.append(f"Row {i}: '{name}' already exists")
            continue

        p = Product(user_id=user.id, name=name, sku=sku or None, category=cat, description=desc)
        db.add(p)
        added.append(name)

    db.commit()

    from app.models import Alert, AlertStatus
    unread_alerts = get_unread_alert_count(user, db)

    return templates.TemplateResponse("bulk_upload.html", {
        "request": request,
        "user": user,
        "unread_alerts": unread_alerts,
        "result": {"added": added, "skipped": skipped, "errors": errors},
    })


@router.post("/products/bulk-upload-files")
async def bulk_upload_files(
    request: Request,
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
):
    """Bulk upload label images/PDFs — creates one product per file and triggers processing."""
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    # Quota gate
    from app.services.quota_service import check_product_limit, get_product_headroom
    from urllib.parse import quote as _quote
    allowed, quota_msg = check_product_limit(user, db)
    if not allowed:
        return RedirectResponse(url=f"/products?msg={_quote(quota_msg)}&type=error", status_code=302)

    # Cap file list to remaining headroom
    headroom = get_product_headroom(user, db)
    cap_warning = None
    if headroom is not None and len(files) > headroom:
        cap_warning = f"Upload capped at {headroom} file(s) — plan limit reached. Upgrade to upload more."
        files = files[:headroom]

    added, skipped, errors = [], [], []
    if cap_warning:
        errors.append(cap_warning)

    for f in files:
        if not f.filename:
            continue
        suffix = Path(f.filename).suffix.lower()
        if suffix not in LABEL_EXTENSIONS:
            errors.append(f"'{f.filename}' — unsupported file type (use JPG, PNG, PDF, WEBP, TIFF)")
            continue

        # Derive product name from filename (strip extension)
        product_name = Path(f.filename).stem.replace("_", " ").replace("-", " ").strip()
        if not product_name:
            product_name = f"Product {uuid.uuid4().hex[:6]}"

        # Skip duplicates
        existing = db.query(Product).filter(
            Product.user_id == user.id,
            Product.name == product_name,
            Product.is_active == True,
        ).first()
        if existing:
            skipped.append(f"'{f.filename}' — product '{product_name}' already exists")
            continue

        try:
            product = Product(
                user_id=user.id,
                name=product_name,
                category="Health Supplement",
            )
            db.add(product)
            db.commit()

            label_version = await _save_label_file(f, suffix, product.id, db)
            background_tasks.add_task(_process_label, label_version.id)
            added.append(product_name)
        except Exception as e:
            errors.append(f"'{f.filename}' — {e}")

    from app.models import Alert, AlertStatus
    unread_alerts = get_unread_alert_count(user, db)

    return templates.TemplateResponse("bulk_upload.html", {
        "request": request,
        "user": user,
        "unread_alerts": unread_alerts,
        "result": {"added": added, "skipped": skipped, "errors": errors},
        "active_tab": "files",
    })


@router.get("/products/archived")
async def archived_products(request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    products = (
        db.query(Product)
        .filter(Product.user_id == user.id, Product.is_active == False)
        .order_by(Product.updated_at.desc())
        .all()
    )
    for p in products:
        _ = p.label_versions
    archived_count = len(products)
    return templates.TemplateResponse("products_archived.html", {
        "request": request,
        "user": user,
        "products": products,
        "archived_count": archived_count,
    })


@router.get("/products/{product_id}")
async def product_detail(product_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    product = (
        db.query(Product)
        .filter(Product.id == product_id, Product.user_id == user.id)
        .options(
            selectinload(Product.label_versions)
            .selectinload(LabelVersion.checks)
            .joinedload(ComplianceCheck.rule)
        )
        .first()
    )
    if not product:
        return RedirectResponse(url="/products")

    return templates.TemplateResponse("product_detail.html", {
        "request": request,
        "user": user,
        "product": product,
        "flash_message": request.query_params.get("msg"),
        "flash_type": request.query_params.get("type", "info"),
    })


@router.post("/products/{product_id}/delete")
async def delete_product(product_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)

    product = db.query(Product).filter(
        Product.id == product_id, Product.user_id == user.id
    ).first()
    if product:
        product.is_active = False
        db.commit()
    return RedirectResponse(url="/products", status_code=302)


@router.post("/products/{product_id}/archive")
async def archive_product(product_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    product = db.query(Product).filter(Product.id == product_id, Product.user_id == user.id).first()
    if product:
        product.is_active = False
        db.commit()
    return RedirectResponse(url="/products", status_code=302)


@router.post("/products/{product_id}/restore")
async def restore_product(product_id: int, request: Request, db: Session = Depends(get_db)):
    user = require_user(request, db)
    if not user:
        return RedirectResponse(url="/login", status_code=302)
    product = db.query(Product).filter(Product.id == product_id, Product.user_id == user.id).first()
    if product:
        product.is_active = True
        db.commit()
    return RedirectResponse(url="/products/archived", status_code=302)


@router.post("/products/bulk-action")
async def bulk_action(request: Request, db: Session = Depends(get_db)):
    """Handle bulk archive or delete for multiple products."""
    from fastapi.responses import JSONResponse as _JSONResponse
    user = require_user(request, db)
    if not user:
        return _JSONResponse({"error": "Not authenticated"}, status_code=401)

    data = await request.json()
    action = data.get("action")  # "archive" | "delete"
    ids = data.get("ids", [])

    if not ids or action not in ("archive", "delete"):
        return _JSONResponse({"error": "Invalid request"}, status_code=400)

    products = db.query(Product).filter(
        Product.id.in_(ids), Product.user_id == user.id
    ).all()

    count = 0
    for p in products:
        if action == "archive":
            p.is_active = False
        elif action == "delete":
            p.is_active = False  # soft-delete only; protects label/check history
        count += 1

    db.commit()
    return _JSONResponse({"success": True, "count": count})
